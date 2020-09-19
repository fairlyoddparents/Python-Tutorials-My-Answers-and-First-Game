import openpyxl

txt = "La comida habla por sí sola, y el menú cuenta con la tradicional bandeja montañera, original de la zona paisa o antioqueña, un plato que desde cualquier punto que se mire es gigantesca con su guarnición de chicharrón piel tostada de cerdo, arroz, frijoles, tajadas de plátano maduro, arepa de maíz, tajada de [Anterior]aguacate[Siguiente], carne molida o frita y huevo frito por encima, es cuestión de gusto y apetito. Algo ya más sofisticado es la cazuela de mariscos, una espesa sopa que hace sudar desde la primera cucharada con un buen aliño marino de camarones, muelas de cangrejo, ostras, anillos y trozos de pulpo."
year = "2003"
author = "PRENSA"
country = "EE. UU."
topic = "05.Gastronomía, cocina"
publication = "Eduardo Marceles (Nueva York), 2003"

class ExcelFile(object):
    book = openpyxl.Workbook()

    def __init__(self, name):
        self.name = name
        """The following code doesn't work! That's why I commented it out!"""
        #sheet = self.book.active
        #sheet['A'].alignment = openpyxl.styles.Alignment(vertical='top', wrap_text=True)
        #self.book.save(f'{self.name}_data.xlsx')

    def add_info(self, info):
        sheet = self.book.active
        sheet.append(info)
        self.book.save(f'{self.name}_data.xlsx')

    def create_sheet(self, sheet_name):
        self.book.create_sheet(sheet_name)
        self.book.save(f'{self.name}_data.xlsx')

    def remove_sheet_by_name(self, sheet_name):
        extra_sheet = self.book[sheet_name]
        self.book.remove(extra_sheet)
        self.book.save(f'{self.name}_data.xlsx')

aguacate_file = ExcelFile('aguacate')
aguacate_file.add_info([txt, year, author, country, topic, publication])
